from square1 import get_data, read_dcm 
import writing2xls_functions as wf
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, numbers, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()
ws1 = wb.active
ws1.title = "All data"
ws2 = wb.create_sheet("Data for sorting")	

ws1.append(["Nr", "Paciento ID", "Tyrimas", "S", "Tyrimo data", "Gimimo metai", "Amžius", "Lytis", "Ūgis, m", "Svoris, kg", "KMI", "Skenavimo vieta", "Protokolas", "Skenavimo ilgis", "n, mm", "N, mm", "p", "U, V", "I, mA","max(I), mA","t, s","total_t, s", "CTDI", "DLP", "Mean CTDI", "Total DLP"])
ws2.append(["Nr", "Paciento ID", "Tyrimas","S", "Tyrimo data", "Gimimo metai", "Amžius", "Lytis", "Ūgis, m", "Svoris, kg", "KMI", "Mean CTDI", "Total DLP"])

Nr = 0
for filename in glob.glob('../input/*.dcm'):
	Nr = Nr + 1
	print (filename, ":")
	extracted_data = get_data(read_dcm(filename))
	n_events = extracted_data['n_events']
	n_scouts = extracted_data['acq_protocol'].count('Topogram')
	n_no_scouts = n_events - n_scouts
	#print (extracted_data['acq_protocol'], n_no_scouts)
	#print(extracted_data)


	# Use "n_events" if we want to include scouts in excel, and "n_no_scouts" if we want to exclude scouts.
	rows_sheet1 = wf.creat_rows_for_sheet1(n_no_scouts, extracted_data, Nr)
	#print (rows_sheet1)
	for row in rows_sheet1:
		ws1.append(row)
	
	rows_sheet2 = wf.creat_rows_for_sheet2(extracted_data, Nr)
	for row in rows_sheet2:
		ws2.append(row)

ws1.insert_cols(idx=1)
ws1.insert_rows(idx=1)
ws2.insert_cols(idx=1)
ws2.insert_rows(idx=1)

for col in ws1.columns:
    column = col[0].column_letter # Get the column name
    adjusted_width = wf.set_auto_column_width(col) 
    ws1.column_dimensions[column].width = adjusted_width

for col in ws2.columns:
    column = col[0].column_letter # Get the column name
    adjusted_width = wf.set_auto_column_width(col) 
    ws2.column_dimensions[column].width = adjusted_width
    
#print (ws1.dimensions, ws2.dimensions) #adding sorting filter to the first row
ws1.auto_filter.ref = "B2:AA10"     
ws2.auto_filter.ref = "B2:N5"

# Let's create a style template for the header row
header = NamedStyle(name="header")
header.font = Font(bold=True)
header.border = Border(bottom=Side(border_style="double"))
header.alignment = Alignment(horizontal="center", vertical="center")

# Now let's apply this to all first row (header) cells
header_row_ws1 = ws1[2]
header_row_ws2 = ws2[2]
for cell in header_row_ws1:
     cell.style = header
for cell in header_row_ws2:
     cell.style = header


wb.save("../output/VULVL_CT_doses_2024.xlsx")











