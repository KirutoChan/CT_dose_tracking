from square1_GE import get_data, read_dcm 
import writing2xls_functions_GE as wf
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()
ws1 = wb.active
ws1.title = "All data"
ws2 = wb.create_sheet("Data for sorting")	

ws1.append(["Nr", "Paciento ID", "Tyrimas", "S", "Tyrimo data", "Gimimo metai", "Amžius", "Lytis", "Ūgis, m", "Svoris, kg", "KMI", "Skenavimo vieta", "Protokolas", "Skenavimo ilgis", "n, mm", "N, mm", "p", "U, V", "I, mA","max(I), mA","t, s","total_t, s", "CTDI", "DLP", "Mean CTDI", "Total DLP"])
ws2.append(["Nr", "Paciento ID", "Tyrimas","S", "Tyrimo data", "Gimimo metai", "Amžius", "Lytis", "Ūgis, m", "Svoris, kg", "KMI", "Mean CTDI", "Total DLP"])

Nr = 0
for filename in glob.glob('../input_GE/*.dcm'):
	Nr = Nr + 1
	print (filename)
	extracted_data = get_data(read_dcm(filename))
	n_events = extracted_data['n_events']
	#print(extracted_data)

	rows_sheet1 = wf.creat_rows_for_sheet1(n_events, extracted_data, Nr)
	for row in rows_sheet1:
		ws1.append(row)
	
	rows_sheet2 = wf.creat_rows_for_sheet2(extracted_data, Nr)
	for row in rows_sheet2:
		ws2.append(row)
	
wb.save("../output_GE/GE.xlsx")


# ws1.column_dimensions["A"].width = 3.0
# ws1.column_dimensions["B"].width = 25.0
# ws1.column_dimensions["C"].width = 30.0
# ws1.column_dimensions["D"].width = 3.0
# ws1.column_dimensions["E"].width = 12.0
# ws1.column_dimensions["F"].width = 14.0
# ws1.column_dimensions["G"].width = 12.0
# ws1.column_dimensions["H"].width = 7.0
# ws1.column_dimensions["I"].width = 5.0
# ws1.column_dimensions["J"].width = 7.0
# ws1.column_dimensions["K"].width = 10.0
# ws1.column_dimensions["L"].width = 5.0
# ws1.column_dimensions["M"].width = 7.0
# ws1.column_dimensions["N"].width = 7.0
# ws1.column_dimensions["O"].width = 7.0
# ws1.column_dimensions["P"].width = 5.0
# ws1.column_dimensions["Q"].width = 6.0
# ws1.column_dimensions["R"].width = 6.0
# ws1.column_dimensions["S"].width = 6.0
# ws1.column_dimensions["T"].width = 7.0
# ws1.column_dimensions["U"].width = 10.0
# ws1.column_dimensions["V"].width = 10.0
# ws2.column_dimensions["A"].width = 3.0
# ws2.column_dimensions["B"].width = 25.0
# ws2.column_dimensions["C"].width = 30.0
# ws2.column_dimensions["D"].width = 3.0
# ws2.column_dimensions["E"].width = 12.0
# ws2.column_dimensions["F"].width = 14.0
# ws2.column_dimensions["G"].width = 12.0
# ws2.column_dimensions["M"].width = 12.0
# ws2.column_dimensions["N"].width = 12.0

# tab = Table(displayName="Table1", ref="A1:N2")

# # Add a default style with striped rows and banded columns
# style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
#                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
# tab.tableStyleInfo = style
# ws2.add_table(tab)










